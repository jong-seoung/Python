from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학 -> 번호 국어 영어 수학 으로 변경
ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어" # B1셀에 국어 입력

ws.move_range("C1:C11",rows=5,cols=-1)
wb.save("sample_korea.xlsx")